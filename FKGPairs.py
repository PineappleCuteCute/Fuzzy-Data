import openpyxl
import numpy as np
import time

# Hàm để nhập dữ liệu từ một sheet trong tệp Excel
def importData(sheet):
    return ([[cell.value for cell in row] for row in sheet])

# Load workbook Excel chứa dữ liệu
wb = openpyxl.load_workbook("Data.xlsx")

def importFuzzyData():# Hàm để nhập dữ liệu mờ từ sheet 'fuzzyData'
    fuzzyData = wb['FuzzyData2']
    base = list(fuzzyData.values)

    return base #trả ra kết quả đầu vào "base"


def combination(k, n):#Tính toán tổ hợp
    if k == 0 or k == n:
        return 1
    if k == 1:
        return n
    return combination(k - 1, n - 1) + combination(k, n - 1)


def caculateA(base):#Tính toán trọng số A dựa trên đầu vào
    colum = len(base[0])
    row = len(base)
    A = np.zeros((row, combination(4, colum - 1)))

    for r1 in range(row):
        k = [0] * combination(4, colum - 1)
        temp = 0
        for a in range(0, colum - 4):
            for b in range(a + 1, colum - 3):
                for c in range(b + 1, colum - 2):
                    for d in range(c + 1, colum - 1):
                        for r2 in range(row):
                            if base[r1][a] == base[r2][a] and base[r1][b] == base[r2][b] and base[r1][c] == base[r2][c] and base[r1][d] == base[r2][d]:
                                k[temp] += 1

                        A[r1][temp] = k[temp] / row
                        temp += 1
    return A


"""A = (caculateA(base))

for i in range(len(A)):
    print(sum(A[i]))"""


def caculateM(base): #Tính toán ma trận M dựa trên đầu vào
    colum = len(base[0])
    row = len(base)
    M = np.zeros((row, colum - 1))
    for t1 in range(row):
        k = [0] * row
        temp = 0
        for i in range(colum - 1):
            for t2 in range(row):
                if base[t1][i] == base[t2][i] and base[t1][colum - 1] == base[t2][colum - 1]:
                    k[temp] += 1
            M[t1][temp] = k[temp] / row
            temp += 1

    return M


# print(caculateM(base))


def caculateB(base, A, M): #Tính toán trọng số B dựa trên trọng số A và ma trận M
    colum = len(base[0])
    row = len(base)
    B = np.zeros((row, combination(3, colum - 1)))

    for r in range(row):
        temp = 0
        for a in range(0, colum - 3):
            for b in range(a + 1, colum - 2):
                for c in range(b + 1, colum - 1):
                    B[r][temp] = sum(A[r]) * min(M[r][a], M[r][b], M[r][c])
                    temp += 1

    return B


# print(caculateB(base, caculateA(base), caculateM(base)))
# print(caculateM(base))


def caculateC(base, B): #Tính trọng số C dựa trên B đã tính ở hàm trên
    colum = len(base[0])
    row = len(base)
    cols = 3 * combination(3, colum - 1)
    C = np.zeros((row, cols))

    for r1 in range(row):
        temp = 0
        for i in range(3):
            for a in range(0, (colum - 3)):
                for b in range(a + 1, (colum - 2)):
                    for c in range(b + 1, (colum - 1)):
                        for r2 in range(row):
                            if base[r1][a] == base[r2][a] and base[r1][b] == base[r2][b] and base[r1][c] == base[r2][c] and base[r2][colum - 1] == i:
                                C[r1][temp] += B[r2][temp % combination(3, colum - 1)]
                        # print(temp,":",temp//combination(3,colum-1))
                        temp += 1

    return C


# print(caculateC(base, caculateB(base, caculateA(base), caculateM(base)))[0])

def writeToExcel(sheet, table, row): #Ghi dữ liệu từ một bảng vào một sheet trong tệp Excel.
    for x in range(row):
        for y in range(len(table[x])):
            sheet.cell(row=x + 1, column=y + 1, value=table[x][y])


def update(): #Cập nhật A, M, B, C và ghi chúng vào các sheet trong workbook của file Execl
    # wb = openpyxl.load_workbook("FKG_Py.xlsx")
    base = importFuzzyData()
    A = caculateA(base)
    M = caculateM(base)
    B = caculateB(base, A, M)
    C = caculateC(base, B)
    writeToExcel(wb['A'], A, len(A))
    writeToExcel(wb['M'], M, len(M))
    writeToExcel(wb['B'], B, len(B))
    writeToExcel(wb['C'], C, len(C))

    wb.save("Data.xlsx")
    print("Update successful! Let's Start")

start = time.time()
update()
print("Time: ", time.time() - start)

def FISA(base, C, list): #Dự đoán đầu ra dựa trên dữ liệu đầu vào và trọng số C.
    colum = len(base[0])
    row = len(base)

    cols = combination(3, (colum - 1))
    C0 = [0] * cols
    C1 = [0] * cols
    C2 = [0] * cols

    t = 0
    for a in range(0, colum - 3):
        for b in range(a + 1, colum - 2):
            for c in range(b + 1, colum - 1):
                for r in range(row-1):
                    if base[r][a] == list[a] and base[r][b] == list[b] and base[r][c] == list[c] and base[r][colum-1] == 0:
                        C0[t] = C[r][t + 0 * cols]
                        break
                    if base[r][a] == list[a] and base[r][b] == list[b] and base[r][c] == list[c] and base[r][colum-1] == 1:
                        C1[t] = C[r][t + 1 * cols]
                        break
                    if base[r][a] == list[a] and base[r][b] == list[b] and base[r][c] == list[c] and base[r][colum-1] == 2:
                        C2[t] = C[r][t + 2 * cols]
                        break
                t += 1
    #print(t)

    D0 = max(C0) + min(C0) #Tính D0
    D1 = max(C1) + min(C1) #Tính D1
    D2 = max(C2) + min(C2) #Tính D2

    #print(D0, max(C0), min(C0))
    #print(D1, max(C1), min(C1))
    #print(D2, max(C2), min(C2))

    if D0 > D1 and D0 > D2: #Dựa vào công thức Max(D0, D1, D2) đưa ra nhãn 0, 1 hoặc 2
        return 0
    elif D1 > D2:
        return 1
    else:
        return 2


def testAccuracy(): #Kiểm tra độ chính xác của mô hình.
    fuzzyTest = wb['FuzzyTest2']
    test = list(fuzzyTest.values)
    sheetC = wb['C']
    base = importFuzzyData()
    #print(base)
    C = list(sheetC.values)
    X = np.zeros((len(test), 1))
    for i in range(len(test)):import openpyxl
import numpy as np
import time

# Hàm để nhập dữ liệu từ một sheet trong tệp Excel
def importData(sheet):
    return ([[cell.value for cell in row] for row in sheet])

# Load workbook Excel chứa dữ liệu
wb = openpyxl.load_workbook("Data.xlsx")

def importFuzzyData():# Hàm để nhập dữ liệu mờ từ sheet 'fuzzyData'
    fuzzyData = wb['FuzzyData2']
    base = list(fuzzyData.values)

    return base #trả ra kết quả đầu vào "base"


def combination(k, n):#Tính toán tổ hợp
    if k == 0 or k == n:
        return 1
    if k == 1:
        return n
    return combination(k - 1, n - 1) + combination(k, n - 1)


def caculateA(base):#Tính toán trọng số A dựa trên đầu vào
    colum = len(base[0])
    row = len(base)
    A = np.zeros((row, combination(4, colum - 1)))

    for r1 in range(row):
        k = [0] * combination(4, colum - 1)
        temp = 0
        for a in range(0, colum - 4):
            for b in range(a + 1, colum - 3):
                for c in range(b + 1, colum - 2):
                    for d in range(c + 1, colum - 1):
                        for r2 in range(row):
                            if base[r1][a] == base[r2][a] and base[r1][b] == base[r2][b] and base[r1][c] == base[r2][c] and base[r1][d] == base[r2][d]:
                                k[temp] += 1

                        A[r1][temp] = k[temp] / row
                        temp += 1
    return A


"""A = (caculateA(base))

for i in range(len(A)):
    print(sum(A[i]))"""


def caculateM(base): #Tính toán ma trận M dựa trên đầu vào
    colum = len(base[0])
    row = len(base)
    M = np.zeros((row, colum - 1))
    for t1 in range(row):
        k = [0] * row
        temp = 0
        for i in range(colum - 1):
            for t2 in range(row):
                if base[t1][i] == base[t2][i] and base[t1][colum - 1] == base[t2][colum - 1]:
                    k[temp] += 1
            M[t1][temp] = k[temp] / row
            temp += 1

    return M


# print(caculateM(base))


def caculateB(base, A, M): #Tính toán trọng số B dựa trên trọng số A và ma trận M
    colum = len(base[0])
    row = len(base)
    B = np.zeros((row, combination(3, colum - 1)))

    for r in range(row):
        temp = 0
        for a in range(0, colum - 3):
            for b in range(a + 1, colum - 2):
                for c in range(b + 1, colum - 1):
                    B[r][temp] = sum(A[r]) * min(M[r][a], M[r][b], M[r][c])
                    temp += 1

    return B


# print(caculateB(base, caculateA(base), caculateM(base)))
# print(caculateM(base))


def caculateC(base, B): #Tính trọng số C dựa trên B đã tính ở hàm trên
    colum = len(base[0])
    row = len(base)
    cols = 3 * combination(3, colum - 1)
    C = np.zeros((row, cols))

    for r1 in range(row):
        temp = 0
        for i in range(3):
            for a in range(0, (colum - 3)):
                for b in range(a + 1, (colum - 2)):
                    for c in range(b + 1, (colum - 1)):
                        for r2 in range(row):
                            if base[r1][a] == base[r2][a] and base[r1][b] == base[r2][b] and base[r1][c] == base[r2][c] and base[r2][colum - 1] == i:
                                C[r1][temp] += B[r2][temp % combination(3, colum - 1)]
                        # print(temp,":",temp//combination(3,colum-1))
                        temp += 1

    return C


# print(caculateC(base, caculateB(base, caculateA(base), caculateM(base)))[0])

def writeToExcel(sheet, table, row): #Ghi dữ liệu từ một bảng vào một sheet trong tệp Excel.
    for x in range(row):
        for y in range(len(table[x])):
            sheet.cell(row=x + 1, column=y + 1, value=table[x][y])


def update(): #Cập nhật A, M, B, C và ghi chúng vào các sheet trong workbook của file Execl
    # wb = openpyxl.load_workbook("FKG_Py.xlsx")
    base = importFuzzyData()
    A = caculateA(base)
    M = caculateM(base)
    B = caculateB(base, A, M)
    C = caculateC(base, B)
    writeToExcel(wb['A'], A, len(A))
    writeToExcel(wb['M'], M, len(M))
    writeToExcel(wb['B'], B, len(B))
    writeToExcel(wb['C'], C, len(C))

    wb.save("Data.xlsx")
    print("Update successful! Let's Start")

start = time.time()
update()
print("Time: ", time.time() - start)

def FISA(base, C, list): #Dự đoán đầu ra dựa trên dữ liệu đầu vào và trọng số C.
    colum = len(base[0])
    row = len(base)

    cols = combination(3, (colum - 1))
    C0 = [0] * cols
    C1 = [0] * cols
    C2 = [0] * cols

    t = 0
    for a in range(0, colum - 3):
        for b in range(a + 1, colum - 2):
            for c in range(b + 1, colum - 1):
                for r in range(row-1):
                    if base[r][a] == list[a] and base[r][b] == list[b] and base[r][c] == list[c] and base[r][colum-1] == 0:
                        C0[t] = C[r][t + 0 * cols]
                        break
                    if base[r][a] == list[a] and base[r][b] == list[b] and base[r][c] == list[c] and base[r][colum-1] == 1:
                        C1[t] = C[r][t + 1 * cols]
                        break
                    if base[r][a] == list[a] and base[r][b] == list[b] and base[r][c] == list[c] and base[r][colum-1] == 2:
                        C2[t] = C[r][t + 2 * cols]
                        break
                t += 1
    #print(t)

    D0 = max(C0) + min(C0) #Tính D0
    D1 = max(C1) + min(C1) #Tính D1
    D2 = max(C2) + min(C2) #Tính D2

    #print(D0, max(C0), min(C0))
    #print(D1, max(C1), min(C1))
    #print(D2, max(C2), min(C2))

    if D0 > D1 and D0 > D2: #Dựa vào công thức Max(D0, D1, D2) đưa ra nhãn 0, 1 hoặc 2
        return 0
    elif D1 > D2:
        return 1
    else:
        return 2


def testAccuracy(): #Kiểm tra độ chính xác của mô hình.
    fuzzyTest = wb['FuzzyTest2']
    test = list(fuzzyTest.values)
    sheetC = wb['C']
    base = importFuzzyData()
    #print(base)
    C = list(sheetC.values)
    X = np.zeros((len(test), 1))
    for i in range(len(test)):
        X[i][0] = FISA(base, C, test[i])
        print(test[i])

    writeToExcel(wb['Result'], X, len(X))
    wb.save("Data.xlsx")

testAccuracy()


"""
sheetC = wb['C']
base = importFuzzyData()
C = importData(sheetC['A1:AFZ139'])
X =
    ['Very high', 'Extremely high', 'Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Low', 'Low','Low', 'High']
print(FISA(base,C, X))
"""


"""
['Very high', 'Very high', 'Medium', 'Medium', 'Medium', 'Medium', 'High', 'Medium', 'Medium', 'Low', 'Low', 'Low', 'High']
['Very high', 'Extremely high', 'Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Low', 'Low', 'Low', 'High']
['Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Low', 'Medium']
['Medium', 'Medium', 'Very high', 'Low', 'Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Low', 'Medium']
['Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'High', 'Medium', 'Medium', 'Medium', 'Medium', 'Low', 'Medium']
['Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Low', 'Medium']
['Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Low', 'Medium']
['Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Low', 'Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Low', 'Medium']
['High', 'Medium', 'High', 'Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Low', 'Medium']
['Very high', 'Very high', 'Low', 'Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Low', 'Low', 'Medium', 'High']
['Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Low', 'Medium']
['Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Low', 'Medium']
['Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Low', 'Medium']
['Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'High', 'Medium', 'Medium', 'Medium', 'Medium', 'Low', 'Medium']
"""
